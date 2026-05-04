#!/usr/bin/env python3
"""Liest KODA-Bot Trades aus Excel und injiziert sie als JSON in bot.html"""
import openpyxl
import json
import os

EXCEL = os.path.join(os.path.dirname(__file__), '..', 'TOCH_Trading_2026_v2.1.xlsx')
BOT_HTML = os.path.join(os.path.dirname(__file__), 'bot.html')

def read_koda_trades():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    if 'KODA-Bot' not in wb.sheetnames:
        print("KODA-Bot Sheet nicht gefunden")
        return []

    ws = wb['KODA-Bot']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    trades = []

    for row in range(2, ws.max_row + 1):
        vals = [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]
        if vals[0] is None:
            continue

        trade = {
            'id': vals[0],
            'datum': str(vals[1]) if vals[1] else '',
            'zeit': str(vals[2]) if vals[2] else '',
            'coin': vals[3] or '',
            'direction': vals[4] or '',
            'strategie': vals[5] or 'ER1',
            'entry': float(vals[6]) if vals[6] else 0,
            'tp1_pct': vals[7] or '',
            'tp1_kurs': float(vals[8]) if vals[8] else None,
            'sl_pct': vals[9] or '',
            'sl_kurs': float(vals[10]) if vals[10] else None,
            'rest_exit': vals[11] or '',
            'rest_kurs': float(vals[12]) if vals[12] else None,
            'roi_pct': float(vals[13]) if vals[13] else 0,
            'pnl_usdt': float(vals[14]) if vals[14] else 0,
            'net_pnl': float(vals[15]) if vals[15] else 0,
            'dauer': str(vals[16]) if vals[16] else '',
            'status': vals[17] or '',
            'ee_main': float(vals[18]) if vals[18] else None,
            'ee_signal': float(vals[19]) if vals[19] else None,
            'steepness': float(vals[20]) if vals[20] else None,
            'golden_wick': vals[21] or '',
            'notiz': vals[22] or '',
        }
        trades.append(trade)

    return trades


def inject_into_html(trades):
    with open(BOT_HTML, 'r') as f:
        html = f.read()

    # Replace or inject KODA_TRADES variable
    json_data = json.dumps(trades, default=str, ensure_ascii=False)
    injection = f"var KODA_TRADES = {json_data};"

    marker = "// Trade data will be injected here by the build script"
    old_line = "// var KODA_TRADES = [{id:1, datum:'2026-05-04', ...}];"

    if marker in html:
        html = html.replace(marker + "\n" + old_line, marker + "\n" + injection)
    elif old_line in html:
        html = html.replace(old_line, injection)
    elif "var KODA_TRADES = " in html:
        # Replace existing injection
        import re
        html = re.sub(r'var KODA_TRADES = .*?;', injection, html)
    else:
        # Append before </script>
        html = html.replace('</script>', injection + '\n</script>')

    with open(BOT_HTML, 'w') as f:
        f.write(html)

    print(f"Injected {len(trades)} trades into bot.html")


if __name__ == '__main__':
    trades = read_koda_trades()
    print(f"Found {len(trades)} KODA-Bot trades")
    inject_into_html(trades)
