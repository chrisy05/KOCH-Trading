#!/usr/bin/env python3
"""
Liest TOCH_Trading_2026_v2.1.xlsx und erzeugt trades_data.js
Alle Dashboard-Seiten (index, trades, stats) laden diese eine Datei.
Aufruf: python3 build_data.py
"""
import json
import openpyxl
import os
from datetime import datetime

EXCEL = os.path.join(os.path.dirname(__file__), "TOCH_Trading_2026_v2.1.xlsx")
OUTPUT = os.path.join(os.path.dirname(__file__), "trades_data.js")

def fmt_duration(val):
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s == "None":
        return ""
    return s

def safe_float(val):
    if val is None:
        return None
    try:
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).replace(",", ".").strip()
        if not s or s == "None" or s == "nicht getradet":
            return None
        return float(s)
    except:
        return None

def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()

def fmt_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%y")
    s = str(val).strip()
    # Normalize date formats
    if "00:00:00" in s:
        s = s.split(" ")[0]
    return s

def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb["Trades"]

    trades = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        nr = row[0]
        if nr is None:
            continue
        try:
            nr = int(nr)
        except:
            continue

        t = {
            "nr": nr,
            "signal": safe_str(row[1]),
            "datum": fmt_date(row[2]),
            "zeit": safe_str(row[3]),
            "account": safe_str(row[4]),
            "coin": safe_str(row[5]),
            "richtung": safe_str(row[6]),
            "kapital": safe_float(row[7]),
            "einsatz": safe_str(row[8]),
            "hebel": safe_str(row[9]),
            "pct_kapital": safe_str(row[10]),
            "geh_einsatz": safe_float(row[11]),
            "entry": safe_str(row[12]),
            "entry2": safe_str(row[13]),
            "entry3": safe_str(row[14]),
            "tp1": safe_str(row[15]),
            "tp2": safe_str(row[16]),
            "tp3": safe_str(row[17]),
            "sl": safe_str(row[18]),
            "ausstieg_dat": fmt_date(row[19]),
            "ausstieg_zeit": safe_str(row[20]),
            "ausstieg": safe_str(row[21]),
            "dauer": fmt_duration(row[22]),
            "roi": safe_float(row[23]),
            "pnl": safe_float(row[24]),
            "net_pnl": safe_float(row[25]),
            "abgeschl": safe_str(row[26]),
            "notiz": safe_str(row[27]) if len(row) > 27 else ""
        }
        trades.append(t)

    # Statistik-Sheet: Einzahlungen und Kontostände
    ws_stat = wb["Statistik"]
    einzahlungen = []
    kontostand = []

    in_einzahlungen = False
    for row in ws_stat.iter_rows(min_row=1, max_row=ws_stat.max_row, values_only=True):
        r0 = safe_str(row[0]) if row[0] else ""
        if "EINZAHLUNGEN" in r0.upper():
            in_einzahlungen = True
            continue
        if "TOTAL" in r0.upper() and in_einzahlungen:
            in_einzahlungen = False
            continue
        if in_einzahlungen and row[0] and row[1]:
            s = safe_float(row[1])
            ini = safe_str(row[2])
            if s and ini in ("TR", "GT"):
                einzahlungen.append({
                    "datum": fmt_date(row[0]),
                    "summe": s,
                    "initialen": ini
                })

    # Trades die nur im Dashboard existieren (nicht in Excel)
    # Diese werden manuell gepflegt bis Excel aktualisiert wird
    EXTRA_FILE = os.path.join(os.path.dirname(__file__), "extra_trades.json")
    if os.path.exists(EXTRA_FILE):
        with open(EXTRA_FILE, "r") as ef:
            extra = json.load(ef)
            trades.extend(extra)
            print(f"  Extra trades loaded: {len(extra)}")

    # --- Payout (Auszahlung) Calculation ---
    PAYOUTS = [
        {"nr": 1, "datum": "31.03.2026", "gewinn": 1090, "ausgaben": 0, "ausgaben_detail": "", "tr": 545, "gt": 0, "ck": 545},
        {"nr": 2, "datum": "07.04.2026", "gewinn": 866, "ausgaben": 0, "ausgaben_detail": "", "tr": 433, "gt": 0, "ck": 433},
        {"nr": 3, "datum": "09.04.2026", "gewinn": 426, "ausgaben": 0, "ausgaben_detail": "", "tr": 213, "gt": 0, "ck": 213},
        {"nr": 4, "datum": "26.04.2026", "gewinn": 2350, "ausgaben": -100, "ausgaben_detail": "Claude Code Abo", "tr": 562.50, "gt": 562.50, "ck": 1125},
        {"nr": 5, "datum": "12.05.2026", "gewinn": 2983, "ausgaben": 0, "ausgaben_detail": "", "tr": 746, "gt": 746, "ck": 1491},
    ]
    LAST_PAYOUT_TRADE_NR = 81

    # Verteilung = gewinn + ausgaben (ausgaben are negative)
    for p in PAYOUTS:
        p["verteilung"] = p["gewinn"] + p["ausgaben"]

    # Current period: trades after last payout
    period_trades = [t for t in trades if t["nr"] > LAST_PAYOUT_TRADE_NR]
    period_pnl = sum(t.get("net_pnl") or t.get("pnl") or 0 for t in period_trades if (t.get("net_pnl") is not None or t.get("pnl") is not None))
    period_count = sum(1 for t in period_trades if t.get("ausstieg_dat"))

    # MT trades (MarkTrade) get 50% of their PnL
    mt_trades = [t for t in period_trades if t.get("signal") == "MarkTrade"]
    mt_pnl_total = sum(t.get("net_pnl") or t.get("pnl") or 0 for t in mt_trades if (t.get("net_pnl") is not None or t.get("pnl") is not None))
    mt_pnl_50 = mt_pnl_total * 0.5

    # Total PnL across ALL trades
    total_pnl = sum(t.get("net_pnl") or t.get("pnl") or 0 for t in trades if (t.get("net_pnl") is not None or t.get("pnl") is not None))

    # Totals from historical payouts
    total_ausbezahlt = sum(p["verteilung"] for p in PAYOUTS)
    total_tr = sum(p["tr"] for p in PAYOUTS)
    total_gt = sum(p["gt"] for p in PAYOUTS)
    total_ck = sum(p["ck"] for p in PAYOUTS)

    payout_data = {
        "payouts": PAYOUTS,
        "last_payout_trade_nr": LAST_PAYOUT_TRADE_NR,
        "einzahlungen": 20000,
        "current_period": {
            "nr": 6,
            "seit": "12.05.2026",
            "trades": period_count,
            "pnl": round(period_pnl, 2),
            "mt_50": round(mt_pnl_50, 2),
            "verteilung": round(period_pnl, 2),
            "tr": round(period_pnl * 0.25, 2),
            "gt": round(period_pnl * 0.25, 2),
            "ck": round(period_pnl * 0.50, 2),
        },
        "totals": {
            "gewinne_gesamt": round(total_pnl, 2),
            "ausbezahlt": round(total_ausbezahlt, 2),
            "tr_kumuliert": round(total_tr + period_pnl * 0.25, 2),
            "gt_kumuliert": round(total_gt + period_pnl * 0.25, 2),
            "ck_kumuliert": round(total_ck + period_pnl * 0.50, 2),
            "offen": round(period_pnl, 2),
        }
    }

    data = {
        "trades": trades,
        "einzahlungen": einzahlungen,
        "generated": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "total_einzahlungen": sum(e["summe"] for e in einzahlungen),
        "payout_data": payout_data
    }

    js = "// Auto-generated from TOCH_Trading_2026_v2.1.xlsx\n"
    js += "// Generated: " + data["generated"] + "\n"
    js += "const TRADE_DATA = " + json.dumps(data, ensure_ascii=False, indent=None) + ";\n"

    with open(OUTPUT, "w", encoding="utf-8") as f:
        f.write(js)

    print(f"Generated {OUTPUT}")
    print(f"  Trades: {len(trades)}")
    print(f"  Einzahlungen: {len(einzahlungen)} ({data['total_einzahlungen']} USDT)")
    print(f"  Stand: {data['generated']}")

if __name__ == "__main__":
    main()
